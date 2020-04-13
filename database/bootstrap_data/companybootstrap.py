#!/usr/bin/env python
# coding: utf-8

import openpyxl 
from datetime import datetime
import psycopg2

# load excel file
path = "./Company Data for Backend Assignment - data_assignment.xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row
max_col = sheet_obj.max_column

fs_map = dict()
fs_map['IPO']='IPO',
fs_map['M&A']='MnA',
fs_map['Seed Round']='SEED_ROUND',
fs_map['Series A']='SERIES_A',
fs_map['Series B']='SERIES_B',
fs_map['Series C']='SERIES_C',
fs_map['Series D']='SERIES_D',
fs_map['Series D+']='SERIES_DPLUS',
fs_map['Unfunded']='UNFUNDED',
fs_map['Unknown']='UNKNOWN'

# read and normalize data
data = list()
for r in range(2, max_row + 1):
    funding_stage = fs_map[sheet_obj.cell(row = r, column = 1).value]
    category = sheet_obj.cell(row = r, column = 2).value
    website = sheet_obj.cell(row = r, column = 3).value
    name = sheet_obj.cell(row = r, column = 4).value
    num_employees = sheet_obj.cell(row = r, column = 5).value
    num_employees_final = None
    if type(num_employees) == str:
        a,b = num_employees.split(',')
        if a == "NA":
            if b == "NA":
                num_employees_final = None
            else:
                num_employees_final = b
        else:
            if b == "NA":
                num_employees_final = a
            else:
                num_employees_final = "{ 0 } - { 1 }".format(a, b)
    elif type(num_employees) == int:
        num_employees_final = str(num_employees)
    vision = sheet_obj.cell(row = r, column = 6).value
    if vision and vision == "nan'":
        vision = None
    data.append((datetime.now(), datetime.now(), name, num_employees_final, vision, website, category, funding_stage))

# insert into database
conn = psycopg2.connect(
    host='localhost',
    database='job_portal',
    user='test',
    password='test',
    port='5432')
query = "INSERT INTO companies (created_at,updated_at,name,num_employees,vision_statement,website,category,funding_stage) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
cur = conn.cursor()
cur.executemany(query, data)
conn.commit()
